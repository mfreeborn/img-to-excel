from functools import partial
from pathlib import Path
from typing import Tuple

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.dimensions import SheetFormatProperties
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image
from progressbar import ProgressBar


def rgb_to_hex(rgb: Tuple[int, int, int]) -> Tuple[str, str, str]:
    """Convert a 3-tuple of decimals representing RGB values to a tuple of hex strings."""
    red = f"{rgb[0]:02X}0000"
    green = f"00{rgb[1]:02X}00"
    blue = f"0000{rgb[2]:02X}"
    return red, green, blue


def get_subpixel_indices(col_num: int) -> Tuple[int, int, int]:
    """Return a 3-tuple of 1-indexed column indices representing subpixels of a single pixel."""
    offset = (col_num - 1) * 2
    red_index = col_num + offset
    green_index = col_num + offset + 1
    blue_index = col_num + offset + 2
    return red_index, blue_index, green_index


def fill_pixel(
    ws: Worksheet,
    row_num: int,
    subpixel_indices: Tuple[int, int, int],
    subpixel_colours: Tuple[str, str, str],
) -> None:
    """Fill the background of a pixel's subpixels with individual colours."""
    pixel_row = partial(ws.cell, row=row_num)
    for subpixel_index, colour in zip(subpixel_indices, subpixel_colours):
        fill_subpixel(pixel_row, subpixel_index, colour)


def fill_subpixel(pixel_row: partial, subpixel_index: int, colour: str) -> None:
    """Fill the background colour of a single subpixel."""
    pixel_row(column=subpixel_index).fill = PatternFill(
        start_color=colour, end_color=colour, fill_type="solid"
    )


def main(input_filepath: Path, destination_filepath: Path) -> None:
    """Take a picture and create a representation of it in Excel by colouring the cells."""
    wb = Workbook()
    ws = wb.active
    # the following appears to equate to 1 pixel * 3 pixels in Excel
    ws.sheet_format = SheetFormatProperties(defaultColWidth=0.1, defaultRowHeight=2.5)

    with Image.open(input_filepath) as im:
        width, height = im.size
        rgb_im = im.convert("RGB")

    with ProgressBar(max_value=width * height) as progress_bar:
        for i, col_num in enumerate(range(1, width + 1)):  # 1-indexed columns
            subpixel_indices = get_subpixel_indices(col_num)

            for row_num in range(1, height + 1):  # 1-indexed rows
                subpixel_colours = rgb_to_hex(rgb_im.getpixel((col_num - 1, row_num - 1)))
                fill_pixel(ws, row_num, subpixel_indices, subpixel_colours)

                iteration = i * height + row_num
                progress_bar.update(iteration)

    print(f"Saving file to {destination_filepath}")
    wb.save(destination_filepath)


if __name__ == "__main__":
    input_filepath = Path("images/simpsons.jpg")
    destination_filepath = Path("output/simpsons.xlsx")

    main(input_filepath, destination_filepath)
